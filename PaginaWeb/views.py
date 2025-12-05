from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, F
from django.contrib import messages
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from django.http import HttpResponse


from .models import (
    UnidadMedida,
    Producto,
    Movimiento,
    Alerta,
    Proveedor,
    Categoria,
    Transportista
)



@login_required
def transportistas(request):
    transportistas = Transportista.objects.filter(activo=True)
    return render(request, "transportistas.html", {"transportistas": transportistas})


@login_required
def transportista_crear(request):
    if request.method == "POST":
        Transportista.objects.create(
            nombre=request.POST.get("nombre"),
            telefono=request.POST.get("telefono"),
            patente=request.POST.get("patente"),
            empresa=request.POST.get("empresa"),
            region=request.POST.get("region"),
            activo=True,
        )
        messages.success(request, "Transportista creado correctamente.")
        return redirect("transportistas")

    return render(request, "transportistas/form.html")

@login_required
def proveedor_editar(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == "POST":
        proveedor.nombre = request.POST.get("nombre")
        proveedor.telefono = request.POST.get("telefono")
        proveedor.email = request.POST.get("email")
        proveedor.direccion = request.POST.get("direccion")
        proveedor.save()

        messages.success(request, "Proveedor actualizado correctamente.")
        return redirect("proveedores")

    return render(request, "proveedores/form.html", {
        "titulo": "Editar Proveedor",
        "proveedor": proveedor
    })


@login_required
def proveedor_eliminar(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    tiene_mov = Movimiento.objects.filter(proveedor_id=id).exists()
    if tiene_mov:
        messages.error(request, "No puedes eliminar este proveedor porque tiene movimientos asociados.")
        return redirect("proveedores")

    proveedor.activo = False
    proveedor.save()
    messages.success(request, "Proveedor eliminado correctamente.")
    return redirect("proveedores")

@login_required
def dashboard(request):
    total_productos = Producto.objects.filter(activo=True).count()
    total_stock = Producto.objects.aggregate(total=Sum("stock_actual"))["total"] or 0
    alertas_activas = Alerta.objects.filter(activa=True).count()

    fecha_inicio = datetime.now() - timedelta(days=30)

    productos_vendidos = (
        Movimiento.objects.filter(tipo="salida", fecha__gte=fecha_inicio)
        .values("producto__nombre")
        .annotate(total=Sum("cantidad"))
        .order_by("-total")[:5]
    )

    movimientos_recientes = (
        Movimiento.objects.select_related("producto", "usuario")
        .order_by("-fecha")[:10]
    )

    movimientos_mes = Movimiento.objects.filter(fecha__gte=fecha_inicio).count()
    alertas_recientes = Alerta.objects.filter(activa=True)[:5]

    return render(request, "dashboard.html", {
        "total_productos": total_productos,
        "total_stock": total_stock,
        "alertas_activas": alertas_activas,
        "movimientos_mes": movimientos_mes,
        "movimientos_recientes": movimientos_recientes,
        "productos_vendidos": productos_vendidos,
        "alertas_recientes": alertas_recientes,
    })



@login_required
def unidades_lista(request):
    unidades = UnidadMedida.objects.all()
    return render(request, "unidades/lista.html", {"unidades": unidades})


@login_required
def unidad_crear(request):
    if request.method == "POST":
        UnidadMedida.objects.create(
            nombre=request.POST.get("nombre"),
            descripcion=request.POST.get("descripcion"),
        )
        messages.success(request, "Unidad creada correctamente.")
        return redirect("unidades_lista")

    return render(request, "unidades/form.html", {"titulo": "Nueva Unidad"})


@login_required
def unidad_editar(request, id):
    unidad = get_object_or_404(UnidadMedida, id=id)

    if request.method == "POST":
        unidad.nombre = request.POST.get("nombre")
        unidad.descripcion = request.POST.get("descripcion")
        unidad.save()
        messages.success(request, "Unidad actualizada correctamente.")
        return redirect("unidades_lista")

    return render(request, "unidades/form.html", {
        "titulo": "Editar Unidad",
        "unidad": unidad
    })


@login_required
def unidad_eliminar(request, id):
    unidad = get_object_or_404(UnidadMedida, id=id)
    unidad.delete()
    messages.success(request, "Unidad eliminada.")
    return redirect("unidades_lista")



@login_required
def productos_lista(request):
    productos = Producto.objects.select_related("unidad_base").filter(activo=True)
    return render(request, "productos/lista.html", {"productos": productos})


@login_required
def producto_crear(request):
    unidades = UnidadMedida.objects.all()

    if request.method == "POST":
        Producto.objects.create(
            codigo=request.POST.get("codigo"),
            nombre=request.POST.get("nombre"),
            unidad_base_id=request.POST.get("unidad_base"),
            stock_minimo=request.POST.get("stock_minimo"),
            descripcion=request.POST.get("descripcion"),
        )
        messages.success(request, "Producto creado correctamente.")
        return redirect("productos_lista")

    return render(request, "productos/form.html", {
        "titulo": "Nuevo Producto",
        "unidades": unidades,
        "producto": None
    })


@login_required
def producto_editar(request, id):
    producto = get_object_or_404(Producto, id=id)
    unidades = UnidadMedida.objects.all()

    if request.method == "POST":
        producto.codigo = request.POST.get("codigo")
        producto.nombre = request.POST.get("nombre")
        producto.unidad_base_id = request.POST.get("unidad_base")
        producto.stock_minimo = request.POST.get("stock_minimo")
        producto.descripcion = request.POST.get("descripcion")
        producto.save()

        messages.success(request, "Producto actualizado correctamente.")
        return redirect("productos_lista")

    return render(request, "productos/form.html", {
        "titulo": "Editar Producto",
        "producto": producto,
        "unidades": unidades
    })

@login_required
def reporte_movimientos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    ws.append(["Fecha", "Producto", "Tipo", "Cantidad", "Usuario"])

    movimientos = Movimiento.objects.all().order_by("-fecha")

    for m in movimientos:
        ws.append([
            m.fecha.strftime("%Y-%m-%d %H:%M"),
            m.producto.nombre if m.producto else "",
            m.tipo,
            float(m.cantidad),
            m.usuario.username if m.usuario else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="movimientos.xlsx"'
    wb.save(response)
    return response

@login_required
def producto_eliminar(request, id):
    producto = get_object_or_404(Producto, id=id)

    if Movimiento.objects.filter(producto=producto).exists():
        messages.error(request, "No puedes eliminar este producto porque tiene movimientos asociados.")
        return redirect("productos_lista")

    producto.delete()
    messages.success(request, "Producto eliminado correctamente.")
    return redirect("productos_lista")



@login_required
def reporte_inventario_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append(["Producto", "Código", "Stock", "Mínimo", "Estado"])

    productos = Producto.objects.filter(activo=True)

    for p in productos:
        ws.append([
            p.nombre,
            p.codigo,
            float(p.stock_actual),
            float(p.stock_minimo),
            p.estado,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="inventario.xlsx"'
    wb.save(response)
    return response
@login_required
def movimientos_lista(request):
    movimientos = Movimiento.objects.select_related("producto", "usuario").all()
    return render(request, "movimientos/lista.html", {"movimientos": movimientos})

@login_required
def proveedor_crear(request):
    if request.method == "POST":
        Proveedor.objects.create(
            nombre=request.POST.get("nombre"),
            telefono=request.POST.get("telefono"),
            email=request.POST.get("email"),
            direccion=request.POST.get("direccion"),
        )
        messages.success(request, "Proveedor registrado correctamente.")
        return redirect("proveedores")

    return render(request, "proveedores/form.html")


@login_required
def reporte_proveedores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    ws.append(["Proveedor", "Producto"])

    productos = Producto.objects.filter(activo=True)

    for p in productos:
        ws.append([
            p.proveedor.nombre if hasattr(p, 'proveedor') else "Sin proveedor",
            p.nombre,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="proveedores.xlsx"'
    wb.save(response)
    return response
@login_required
def movimiento_crear(request):
    productos = Producto.objects.filter(activo=True)
    proveedores = Proveedor.objects.filter(activo=True)

    if request.method == "POST":

        # Validaciones básicas
        producto_id = request.POST.get("producto")
        tipo = request.POST.get("tipo")
        proveedor_id = request.POST.get("proveedor") or None
        motivo = request.POST.get("motivo")
        cantidad_raw = request.POST.get("cantidad")

        if not producto_id:
            messages.error(request, "Debe seleccionar un producto.")
            return redirect("movimiento_crear")

        # Proveedor obligatorio solo en entrada
        if tipo == "entrada" and not proveedor_id:
            messages.error(request, "Debe seleccionar un proveedor.")
            return redirect("movimiento_crear")

        # Parse cantidad
        try:
            cantidad = Decimal(cantidad_raw)
        except:
            messages.error(request, "Cantidad inválida.")
            return redirect("movimiento_crear")

        producto_obj = Producto.objects.get(id=producto_id)

        # No permitir salidas sin stock
        if tipo in ["salida", "merma"] and cantidad > producto_obj.stock_actual:
            messages.error(request, "Stock insuficiente para realizar la operación.")
            return redirect("movimiento_crear")

        # Crear movimiento
        Movimiento.objects.create(
            producto=producto_obj,
            tipo=tipo,
            unidad_ingreso=producto_obj.unidad_base,
            cantidad=cantidad,
            factor=Decimal(1),
            proveedor_id=proveedor_id,
            motivo=motivo,
            usuario=request.user,
        )

        # Actualizar stock
        if tipo == "entrada":
            producto_obj.stock_actual += cantidad
        else:
            producto_obj.stock_actual -= cantidad

        producto_obj.save()

        # Alertas
        if producto_obj.stock_actual < producto_obj.stock_minimo:
            Alerta.objects.update_or_create(
                producto=producto_obj,
                defaults={"activa": True}
            )
        else:
            Alerta.objects.filter(producto=producto_obj).update(activa=False)

        messages.success(request, "Movimiento registrado correctamente.")
        return redirect("movimientos_lista")

    return render(request, "movimientos/form.html", {
        "productos": productos,
        "proveedores": proveedores
    })




@login_required
def inventario(request):
    categoria_id = request.GET.get("categoria")
    estado = request.GET.get("estado")
    busqueda = request.GET.get("busqueda")

    productos = Producto.objects.filter(activo=True).select_related("unidad_base")

    # Filtrar categoría
    if categoria_id and categoria_id.isdigit():
        productos = productos.filter(categoria_id=categoria_id)

    # Filtros por estado
    if estado == "agotado":
        productos = productos.filter(stock_actual__lte=0)

    elif estado == "critico":
        productos = productos.filter(stock_actual__lt=F("stock_minimo"))

    elif estado == "bajo":
        productos = productos.filter(
            stock_actual__gte=F("stock_minimo"),
            stock_actual__lt=F("stock_minimo") * 1.5
        )

    elif estado == "sobrestock":
        productos = productos.filter(stock_actual__gte=F("stock_minimo") * 2)

    # Búsqueda
    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda)
        )

    return render(request, "inventario.html", {
        "productos": productos.order_by("nombre"),
        "categorias": Categoria.objects.all(),
        "categoria_seleccionada": categoria_id,
        "estado_seleccionado": estado,
        "busqueda": busqueda,
        "total_stock": productos.aggregate(total=Sum("stock_actual"))["total"] or 0,
        "criticos": productos.filter(stock_actual__lt=F("stock_minimo")).count(),
        "agotados": productos.filter(stock_actual__lte=0).count(),
    })



@login_required
def proveedores(request):
    proveedores = Proveedor.objects.filter(activo=True)
    return render(request, "proveedores.html", {"proveedores": proveedores})


@login_required
def reportes(request):
    return render(request, "reportes.html")


@login_required
def alertas(request):
    alertas = Alerta.objects.filter(activa=True).select_related("producto")
    return render(request, "alertas.html", {"alertas": alertas})
